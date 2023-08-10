import math
import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from getter import get_wafer, get_R_structures, get_sessions, get_structures, get_coords, get_Leak_structures, \
    get_C_structures, get_Cmes_structures, get_map_structures
from scipy.stats import percentileofscore
import numpy as np


def wanted_excel(wafer_id, sessions, structures, types, Temps, Files, coords, file_name):
    """
    This function creates an Excel file with given information and register it in a folder named following the concerned wafer.
    We first get all information from the database into a Pandas' DataFrame, and then we write the DataFrame into an Excel.
    One column is: [name of the session, name of the structure, Unit of the Measure, [Measures]]
    One sheet is created per die and one file is created per type of measure
    Size of columns are automatically adjusted for better readability.

    :param wafer_id: ID of the Wafer
    :param sessions: All sessions that we want to write
    :param structures: All structures that we want to write
    :param types: All types that we want to write
    :param Temps: All temperatures that we want to write
    :param Files: All files that we want to write
    :param coords: All coordinates that we want to write
    :param file_name: Name under which the file will be registered

    """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)
    wafer = get_wafer(wafer_id)
    list_of_sheets = []
    df_dict = {}

    df_dict["I"] = {}
    df_dict["J"] = {}
    df_dict["C"] = {}
    df_dict["It"] = {}

    for session in sessions:
        for structure_id in structures:
            if wafer[session].get(structure_id) is None:
                continue
            for matrix in wafer[session][structure_id]["matrices"]:
                coord = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'
                if coord in coords:
                    if coord not in df_dict.keys():
                        df_dict[coord] = pd.DataFrame()
                    for type in matrix["results"]:
                        columns = [
                                      f"{structure_id} {' '.join(session.split(' ')[1:])} {coord} {matrix['results'][type]['Temperature']} {matrix['results'][type]['Filename']}"] + [
                                      f"{structure_id} {' '.join(session.split(' ')[1:])} {coord} {matrix['results'][type]['Temperature']} {matrix['results'][type]['Filename']}" + " "]

                        if matrix["results"][type]["Filename"] in Files and matrix["results"][type][
                            "Temperature"] in Temps:
                            if type == "I" and 'I' in types:
                                if os.path.exists(f"{wafer_id}\\{file_name}_I-V.xlsx"):
                                    continue
                                if coord not in df_dict["I"]:
                                    df_dict["I"][coord] = pd.DataFrame()

                                voltages = [session, matrix["results"][type]["Temperature"],
                                            matrix["results"][type]["Filename"], 'Voltage (V)']
                                I = [session, matrix["results"][type]["Temperature"],
                                     matrix["results"][type]["Filename"], 'I (A)']
                                for double in matrix["results"][type]["Values"]:
                                    voltages.append(double["V"])
                                    I.append(double["I"])

                                new_df = pd.DataFrame(list(zip(voltages, I)), columns=columns)
                                df_dict["I"][coord] = df_dict["I"][coord].merge(new_df, left_index=True,
                                                                                right_index=True,
                                                                                how='outer')
                                I.clear()
                                voltages.clear()

                            elif type == "J" and 'J' in types:
                                if coord not in df_dict["J"]:
                                    df_dict["J"][coord] = pd.DataFrame()
                                if os.path.exists(f"{wafer_id}\\{file_name}_J-V.xlsx"):
                                    continue
                                voltages = [session, matrix["results"][type]["Temperature"],
                                            matrix["results"][type]["Filename"], 'Voltage (V)']
                                J = [session, matrix["results"][type]["Temperature"],
                                     matrix["results"][type]["Filename"], 'J (A/cm^2)']
                                for double in matrix["results"][type]["Values"]:

                                    voltages.append(double["V"])
                                    J.append(double["J"])


                                new_df = pd.DataFrame(list(zip(voltages, J)), columns=columns)
                                df_dict["J"][coord] = df_dict["J"][coord].merge(new_df, left_index=True,
                                                                                right_index=True,
                                                                                how='outer')
                                J.clear()
                                voltages.clear()

                            elif type == "C" and 'C' in types:
                                if os.path.exists(f"{wafer_id}\\{file_name}_C-V.xlsx"):
                                    continue

                                if coord not in df_dict["C"]:
                                    df_dict["C"][coord] = pd.DataFrame()
                                voltages = [session, matrix["results"][type]["Temperature"],
                                            matrix["results"][type]["Filename"], 'Voltage (V)']

                                CS = [session, matrix["results"][type]["Temperature"],
                                      matrix["results"][type]["Filename"], 'CS (Ω)']

                                RS = [session, matrix["results"][type]["Temperature"],
                                      matrix["results"][type]["Filename"], 'RS (Ω)']

                                for double in matrix["results"][type]["Values"]:
                                    voltages.append(double["V"])
                                    CS.append(double["CS"])
                                    RS.append(double["RS"])

                                new_df = pd.DataFrame(list(zip(voltages, CS, RS)),
                                                      columns=columns + [f"{structure_id} {' '.join(session.split(' ')[1:])} {coord} {matrix['results'][type]['Temperature']} {matrix['results'][type]['Filename']}" + " " + " "])
                                df_dict["C"][coord] = df_dict["C"][coord].merge(new_df, left_index=True,
                                                                                right_index=True,
                                                                                how='outer')
                                voltages.clear()
                                CS.clear()
                                RS.clear()

                            elif type == "It" and 'It' in types:
                                if os.path.exists(f"{wafer_id}\\{file_name}_It-V.xlsx"):
                                    continue
                                if coord not in df_dict["It"]:
                                    df_dict["It"][coord] = pd.DataFrame()
                                voltages = [session, matrix["results"][type]["Temperature"],
                                            matrix["results"][type]["Filename"], 'Voltage (V)']
                                It = [session, matrix["results"][type]["Temperature"],
                                      matrix["results"][type]["Filename"], 'TDDB']
                                for double in matrix["results"][type]["Values"]:
                                    voltages.append(double["V"])
                                    It.append(double["TDDB"])

                                new_df = pd.DataFrame(list(zip(voltages, It)), columns=columns)
                                df_dict["It"][coord] = df_dict["It"][coord].merge(new_df, left_index=True,
                                                                                  right_index=True,
                                                                                  how='outer')
                                voltages.clear()
                                It.clear()

    for element in types:
        if df_dict[element] != {}:
            filename = wafer_id + '\\' + file_name + "_" + element + '-V.xlsx'

            pd.DataFrame().to_excel(filename, index=False)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for coord, df in df_dict[element].items():
                    # Créez une copie de df et modifiez les noms de colonnes de la copie
                    df_copy = df.copy()
                    df_copy.columns = [col.split(' ')[0] for col in df_copy.columns]

                    sheetName = coord
                    if sheetName in list_of_sheets:
                        df_copy.to_excel(writer, sheet_name=sheetName, index=False,
                                         startcol=writer.sheets[sheetName].max_column)
                    else:
                        df_copy.to_excel(writer, sheet_name=sheetName, index=False)
                    list_of_sheets.append(sheetName)
                df_dict[element].clear()
            list_of_sheets.clear()

            wb = load_workbook(filename)
            if 'Sheet1' in wb.sheetnames:
                del wb['Sheet1']
            wb.save(filename)


def excel_VBD(wafer_id, sessions, structures, Temps, Files, coords, file_name):
    """
    An Excel file is created with all VBDs inside the wafer, following selected filters.
    3 sheets are created: one for positive values, one for negatives and one for NaN

    :param wafer_id: ID of the wafer
    :param sessions: Filtered sessions
    :param structures: Filtered structures
    :param Temps: Filtered temperatures
    :param Files: Filtered files
    :param coords: Filtered coordinates
    :param file_name: Name under which the file will be registered

    """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    file_name = wafer_id + '\\' + file_name + '.xlsx'
    wafer = get_wafer(wafer_id)

    final_df = pd.DataFrame(index=list(set(structures)))

    for session in sessions:
        temporary_df = pd.DataFrame(index=structures, columns=pd.MultiIndex.from_product([[session], coords]))
        for structure in structures:
            for matrix in wafer[session][structure]["matrices"]:
                coordinates = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'
                if coordinates in coords and matrix["results"]["I"]["Filename"] in Files and matrix["results"]["I"][
                            "Temperature"] in Temps:
                    temporary_df.loc[structure, (session, coordinates)] = matrix['VBD'] if not math.isnan(matrix['VBD']) else 'NaN'
        final_df = pd.concat([final_df, temporary_df], axis=1)

    final_df.to_excel(file_name, sheet_name=wafer_id, index=structures)
    final_df['Classification'] = final_df.apply(classify_row, axis=1)

# Create separated DataFrames
    positive_df = final_df[final_df['Classification'] == 'Positives'].drop(columns='Classification')
    negative_df = final_df[final_df['Classification'] == 'Negatives'].drop(columns='Classification')
    nan_df = final_df[final_df['Classification'] == 'NaN'].drop(columns='Classification')

    # Write each DataFrame to a separate sheet
    with pd.ExcelWriter(file_name) as writer:
        if not positive_df.empty:
            positive_df.to_excel(writer, sheet_name='Positives')
        if not negative_df.empty:
            negative_df.to_excel(writer, sheet_name='Negatives')
        if not nan_df.empty:
            nan_df.to_excel(writer, sheet_name='NaN')

    # Charger le classeur et ajuster la largeur des colonnes
    wb = load_workbook(file_name)

    for sheet in wb.sheetnames:
        wb[sheet].insert_rows(1)

        header_1 = [col[0] for col in final_df.columns]
        header_2 = [col[1] for col in final_df.columns]

        for i, headers in enumerate(zip(header_1, header_2), start=2):
            wb[sheet].cell(row=1, column=i, value=headers[0])
            wb[sheet].cell(row=2, column=i, value=headers[1])

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(file_name)


def classify_row(row):
    """
    Used to know in which sheet the corresponding row has to go

    :param row: Row of the DataFrame

    :return: A str : 'Positives', 'Negatives' or 'NaN'
    """
    for value in row.values:
        numeric_value = pd.to_numeric(value, errors='coerce')
        if pd.notna(numeric_value):
            return 'Positives' if numeric_value > 0 else 'Negatives'
    return 'NaN'


def excel_normal_R(wafer_id, sessions, structures, coords, file_name):
    """
    Used to create an Excel file with all selected values of Resistance and their percentiles of normal distribution.
    Filename has to be entered without any extension.

    :param wafer_id: ID of the wafer
    :param sessions: All sessions to be processed
    :param structures: All structures to be processed
    :param coords: All dies to be processed
    :param file_name: name of the file created
    """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    big_pos_df = pd.DataFrame()
    big_neg_df = pd.DataFrame()

    wafer = get_wafer(wafer_id)
    for session in sessions:
        pos_values = []
        neg_values = []
        for structure in list(set(structures) & set(get_R_structures(wafer_id, session))):
            for matrix in wafer[session][structure]['matrices']:
                coord = f"({matrix['coordinates']['x']},{matrix['coordinates']['y']})"
                if matrix.get("R") is not None and coord in coords:
                    if matrix['R'] >= 0:
                        if matrix['R'] < 999000:
                            pos_values.append(matrix['R'])
                        else:
                            pos_values.append(1e30)
                    else:
                        neg_values.append(matrix['R'])

        pos_values.sort()
        neg_values.sort()

        percentiles = ['Percentiles']

        pos_perc = []
        n = len(pos_values)
        for i in range(0, n):
            pos_perc.append(100*((i+1-0.5)/n))
        pos_percentiles = percentiles + pos_perc

        neg_perc = []
        n = len(neg_values)
        for i in range(0, n):
            neg_perc.append(100 * ((i + 1 - 0.5) / n))
        neg_percentiles = percentiles + neg_perc

        pos_df = pd.DataFrame({
            session: ['Resistance'] + pos_values,
            session + ' ': pos_percentiles
        })

        neg_df = pd.DataFrame({
            session: ['Resistance'] + neg_values,
            session + ' ': neg_percentiles
        })

        big_pos_df = big_pos_df.merge(pos_df, left_index=True, right_index=True, how='outer')
        big_neg_df = big_neg_df.merge(neg_df, left_index=True, right_index=True, how='outer')

    with pd.ExcelWriter(f"{wafer_id}\\{file_name}.xlsx", engine='openpyxl') as writer:
        if not big_pos_df.empty:
            big_pos_df.to_excel(writer, sheet_name="Positives", index=False)
        if not big_neg_df.empty:
            big_neg_df.to_excel(writer, sheet_name="Negatives", index=False)

    wb = load_workbook(f"{wafer_id}\\{file_name}.xlsx")

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(f"{wafer_id}\\{file_name}.xlsx")


def excel_normal_Leak(wafer_id, sessions, structures, coords, file_name):
    """
        Used to create an Excel file with all selected values of Leakage and their percentiles of normal distribution.
        Filename has to be entered without any extension.

        :param wafer_id: ID of the wafer
        :param sessions: All sessions to be processed
        :param structures: All structures to be processed
        :param coords: All dies to be processed
        :param file_name: name of the file created
        """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    big_df = pd.DataFrame()

    wafer = get_wafer(wafer_id)
    for session in sessions:
        values = []
        for structure in list(set(structures) & set(get_Leak_structures(wafer_id, session))):
            for matrix in wafer[session][structure]['matrices']:
                coord = f"({matrix['coordinates']['x']},{matrix['coordinates']['y']})"
                if matrix.get("Leak") is not None and coord in coords:
                    values.append(matrix['Leak'])


        values.sort()
        percentiles = ['Percentiles']
        perc = []
        n = len(values)
        for i in range(0, n):
            perc.append(100 * ((i + 1 - 0.5) / n))
        percentiles = percentiles + perc

        df = pd.DataFrame({
            session: ['Leakage'] + values,
            session + ' ': percentiles
        })

        big_df = big_df.merge(df, left_index=True, right_index=True, how='outer')

    big_df.to_excel(f"{wafer_id}\\{file_name}.xlsx", index=False, engine='openpyxl')

    wb = load_workbook(f"{wafer_id}\\{file_name}.xlsx")

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(f"{wafer_id}\\{file_name}.xlsx")


def excel_normal_C(wafer_id, sessions, structures, coords, file_name):
    """
        Used to create an Excel file with all selected values of Capacitance and their percentiles of normal distribution.
        Filename has to be entered without any extension.

        :param wafer_id: ID of the wafer
        :param sessions: All sessions to be processed
        :param structures: All structures to be processed
        :param coords: All dies to be processed
        :param file_name: name of the file created
        """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    big_df = pd.DataFrame()

    wafer = get_wafer(wafer_id)
    for session in sessions:
        values = []
        for structure in list(set(structures) & set(get_C_structures(wafer_id, session))):
            for matrix in wafer[session][structure]['matrices']:
                coord = f"({matrix['coordinates']['x']},{matrix['coordinates']['y']})"
                if matrix.get("Cap") is not None and coord in coords:
                    if matrix["Cap"].get("C"):
                        values.append(matrix['Cap']['C'])


        values.sort()
        percentiles = ['Percentiles']
        perc = []
        n = len(values)
        for i in range(0, n):
            perc.append(100 * ((i + 1 - 0.5) / n))
        percentiles = percentiles + perc

        df = pd.DataFrame({
            session: ['Capacitance'] + values,
            session + ' ': percentiles
        })

        big_df = big_df.merge(df, left_index=True, right_index=True, how='outer')

    big_df.to_excel(f"{wafer_id}\\{file_name}.xlsx", index=False, engine='openpyxl')

    wb = load_workbook(f"{wafer_id}\\{file_name}.xlsx")

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(f"{wafer_id}\\{file_name}.xlsx")


def excel_normal_Cmes(wafer_id, sessions, structures, coords, file_name):
    """
        Used to create an Excel file with all selected values of Measured Capacitance and their percentiles of normal distribution.
        Filename has to be entered without any extension.

        :param wafer_id: ID of the wafer
        :param sessions: All sessions to be processed
        :param structures: All structures to be processed
        :param coords: All dies to be processed
        :param file_name: name of the file created
        """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    big_df = pd.DataFrame()

    wafer = get_wafer(wafer_id)
    for session in sessions:
        values = []
        for structure in list(set(structures) & set(get_Cmes_structures(wafer_id, session))):
            for matrix in wafer[session][structure]['matrices']:
                coord = f"({matrix['coordinates']['x']},{matrix['coordinates']['y']})"
                if matrix.get("Cap") is not None and coord in coords:
                    if matrix["Cap"].get("Cmes"):
                        values.append(matrix['Cap']['Cmes'])


        values.sort()
        percentiles = ['Percentiles']
        perc = []
        n = len(values)
        for i in range(0, n):
            perc.append(100 * ((i + 1 - 0.5) / n))
        percentiles = percentiles + perc

        df = pd.DataFrame({
            session: ['Measured Capacitance'] + values,
            session + ' ': percentiles
        })

        big_df = big_df.merge(df, left_index=True, right_index=True, how='outer')

    big_df.to_excel(f"{wafer_id}\\{file_name}.xlsx", index=False, engine='openpyxl')

    wb = load_workbook(f"{wafer_id}\\{file_name}.xlsx")

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(f"{wafer_id}\\{file_name}.xlsx")


def excel_normal_VBD(wafer_id, sessions, structures, coords, file_name):
    """
        Used to create an Excel file with all selected values of VBD and their percentiles of normal distribution.
        Filename has to be entered without any extension.

        :param wafer_id: ID of the wafer
        :param sessions: All sessions to be processed
        :param structures: All structures to be processed
        :param coords: All dies to be processed
        :param file_name: name of the file created
        """
    if not os.path.exists(wafer_id):
        os.makedirs(wafer_id)

    big_pos_df = pd.DataFrame()
    big_neg_df = pd.DataFrame()

    wafer = get_wafer(wafer_id)
    for session in sessions:
        pos_values = []
        neg_values = []
        for structure in list(set(structures) & set(get_R_structures(wafer_id, session))):
            for matrix in wafer[session][structure]['matrices']:
                coord = f"({matrix['coordinates']['x']},{matrix['coordinates']['y']})"
                if matrix.get("VBD") is not None and coord in coords and not np.isnan(matrix["VBD"]):
                    if matrix['VBD'] >= 0:
                        pos_values.append(matrix['VBD'])
                    else:
                        neg_values.append(matrix['VBD'])

        pos_values.sort()
        neg_values.sort()

        percentiles = ['Percentiles']

        pos_perc = []
        n = len(pos_values)
        for i in range(0, n):
            pos_perc.append(100 * ((i + 1 - 0.5) / n))
        pos_percentiles = percentiles + pos_perc

        neg_perc = []
        n = len(neg_values)
        for i in range(0, n):
            neg_perc.append(100 * ((i + 1 - 0.5) / n))
        neg_percentiles = percentiles + neg_perc

        pos_df = pd.DataFrame({
            session: ['VBD'] + pos_values,
            session + ' ': pos_percentiles
        })

        neg_df = pd.DataFrame({
            session: ['VBD'] + neg_values,
            session + ' ': neg_percentiles
        })

        big_pos_df = big_pos_df.merge(pos_df, left_index=True, right_index=True, how='outer')
        big_neg_df = big_neg_df.merge(neg_df, left_index=True, right_index=True, how='outer')

    with pd.ExcelWriter(f"{wafer_id}\\{file_name}.xlsx", engine='openpyxl') as writer:
        if not big_pos_df.empty:
            big_pos_df.to_excel(writer, sheet_name="Positives", index=False)
        if not big_neg_df.empty:
            big_neg_df.to_excel(writer, sheet_name="Negatives", index=False)

    wb = load_workbook(f"{wafer_id}\\{file_name}.xlsx")

    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(f"{wafer_id}\\{file_name}.xlsx")


